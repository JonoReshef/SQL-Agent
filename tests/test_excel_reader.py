"""Tests for Excel report ingestion"""

import sys
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

import pytest
from openpyxl import Workbook
from src.ingestion.excel_reader import (
    _parse_properties_string,
    load_report,
    read_email_summary,
    read_product_mentions,
)


@pytest.fixture
def sample_report_excel(tmp_path):
    """Create a sample report Excel file for testing"""
    excel_path = tmp_path / "test_report.xlsx"

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Create Product Mentions sheet
    ws_mentions = wb.create_sheet("Product Mentions")
    ws_mentions.append(
        [
            "Email information",
            "Category",
            "Properties",
            "Requestor",
            "Quantity",
            "Unit",
            "Context",
            "Date Requested",
            "Email Subject",
            "Sender",
            "File",
        ]
    )
    ws_mentions.append(
        [
            "100 pcs of 1/2-13 Grade 8 hex bolts",
            "Fasteners",
            "grade=8, size=1/2-13",
            "john.doe@example.com",
            100,
            "pcs",
            "quote_request",
            "2025-02-15",
            "Request for bolts",
            "customer@example.com",
            "email1.msg",
        ]
    )
    ws_mentions.append(
        [
            "M12 x 50mm stainless steel bolts",
            "Fasteners",
            "size=M12, length=50mm, material=stainless steel",
            "jane.smith@example.com",
            50,
            "pcs",
            "order",
            "2025-02-16",
            "Order confirmation",
            "supplier@example.com",
            "email2.msg",
        ]
    )

    # Create Email Summary sheet
    ws_summary = wb.create_sheet("Email Summary")
    ws_summary.append(["Subject", "Sender", "Date", "Products Count"])
    ws_summary.append(["Request for bolts", "customer@example.com", "2025-02-15", 1])
    ws_summary.append(["Order confirmation", "supplier@example.com", "2025-02-16", 1])

    wb.save(excel_path)
    wb.close()

    return excel_path


@pytest.fixture
def missing_sheet_excel(tmp_path):
    """Create an Excel file missing the Product Mentions sheet"""
    excel_path = tmp_path / "missing_sheet.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Wrong Sheet"
    ws.append(["Data"])

    wb.save(excel_path)
    wb.close()

    return excel_path


@pytest.fixture
def missing_columns_excel(tmp_path):
    """Create an Excel file with missing required columns"""
    excel_path = tmp_path / "missing_columns.xlsx"

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    ws = wb.create_sheet("Product Mentions")
    ws.append(["Email information", "Properties"])  # Missing Category, etc.
    ws.append(["Test product", "grade=8"])

    wb.save(excel_path)
    wb.close()

    return excel_path


@pytest.fixture
def empty_rows_excel(tmp_path):
    """Create an Excel file with empty rows"""
    excel_path = tmp_path / "empty_rows.xlsx"

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    ws = wb.create_sheet("Product Mentions")
    ws.append(
        [
            "Email information",
            "Category",
            "Properties",
            "Requestor",
            "Quantity",
            "Unit",
            "Context",
            "Date Requested",
            "Email Subject",
            "Sender",
            "File",
        ]
    )
    ws.append(
        [
            "Valid product",
            "Fasteners",
            "grade=8",
            "test@example.com",
            100,
            "pcs",
            "quote",
            "2025-02-15",
            "Test",
            "sender@example.com",
            "test.msg",
        ]
    )
    ws.append(
        [None, None, None, None, None, None, None, None, None, None, None]
    )  # Empty row
    ws.append(
        [
            "",
            "",
            "",
            "",
            None,
            None,
            None,
            None,
            "",
            "",
            None,
        ]
    )  # Another empty row
    ws.append(
        [
            "Another valid product",
            "Hardware",
            "size=M10",
            "test2@example.com",
            50,
            "pcs",
            "order",
            "2025-02-16",
            "Order",
            "sender2@example.com",
            "test2.msg",
        ]
    )

    wb.save(excel_path)
    wb.close()

    return excel_path


def test_read_product_mentions_success(sample_report_excel):
    """Test reading product mentions from valid report"""
    mentions = read_product_mentions(sample_report_excel)

    assert len(mentions) == 2

    # Check first mention
    assert mentions[0].exact_product_text == "100 pcs of 1/2-13 Grade 8 hex bolts"
    assert mentions[0].product_category == "Fasteners"
    assert mentions[0].quantity == 100
    assert mentions[0].unit == "pcs"
    assert mentions[0].context == "quote_request"
    assert mentions[0].email_subject == "Request for bolts"
    assert mentions[0].email_sender == "customer@example.com"

    # Check properties
    assert len(mentions[0].properties) == 2
    assert mentions[0].properties[0].name == "grade"
    assert mentions[0].properties[0].value == "8"
    assert mentions[0].properties[1].name == "size"
    assert mentions[0].properties[1].value == "1/2-13"


def test_read_product_mentions_file_not_found():
    """Test error when file doesn't exist"""
    with pytest.raises(FileNotFoundError, match="Report file not found"):
        read_product_mentions("nonexistent.xlsx")


def test_read_product_mentions_missing_sheet(missing_sheet_excel):
    """Test error when Product Mentions sheet is missing"""
    with pytest.raises(ValueError, match="Sheet 'Product Mentions' not found"):
        read_product_mentions(missing_sheet_excel)


def test_read_product_mentions_missing_columns(missing_columns_excel):
    """Test error when required columns are missing"""
    with pytest.raises(ValueError, match="Missing required columns"):
        read_product_mentions(missing_columns_excel)


def test_read_product_mentions_with_empty_rows(empty_rows_excel):
    """Test skipping empty rows"""
    mentions = read_product_mentions(empty_rows_excel)

    assert len(mentions) == 2  # Should skip the 2 empty rows
    assert mentions[0].exact_product_text == "Valid product"
    assert mentions[1].exact_product_text == "Another valid product"


def test_read_email_summary_success(sample_report_excel):
    """Test reading email summary from valid report"""
    summary = read_email_summary(sample_report_excel)

    assert len(summary) == 2
    assert summary[0]["Subject"] == "Request for bolts"
    assert summary[0]["Sender"] == "customer@example.com"
    assert summary[1]["Subject"] == "Order confirmation"


def test_read_email_summary_file_not_found():
    """Test error when file doesn't exist"""
    with pytest.raises(FileNotFoundError, match="Report file not found"):
        read_email_summary("nonexistent.xlsx")


def test_load_report_success(sample_report_excel):
    """Test loading complete report"""
    report = load_report(sample_report_excel)

    assert "product_mentions" in report
    assert "email_summary" in report
    assert "report_path" in report
    assert "total_products" in report
    assert "total_emails" in report

    assert report["total_products"] == 2
    assert report["total_emails"] == 2
    assert len(report["product_mentions"]) == 2
    assert len(report["email_summary"]) == 2


def test_parse_properties_string():
    """Test parsing properties string"""
    # Valid properties
    props = _parse_properties_string("grade=8, size=1/2-13, material=steel")
    assert len(props) == 3
    assert props[0].name == "grade"
    assert props[0].value == "8"
    assert props[1].name == "size"
    assert props[1].value == "1/2-13"
    assert props[2].name == "material"
    assert props[2].value == "steel"

    # Empty string
    assert _parse_properties_string("") == []
    assert _parse_properties_string("nan") == []
    assert _parse_properties_string("none") == []

    # Malformed strings
    assert _parse_properties_string("no equals sign") == []
    assert _parse_properties_string("grade=") == []  # Empty value
    assert _parse_properties_string("=8") == []  # Empty name


@pytest.mark.integration
def test_read_real_report_if_exists():
    """Test reading a real report file if it exists"""
    report_path = Path("output/report_20251114_123659.xlsx")

    if not report_path.exists():
        pytest.skip(f"Real report file not found: {report_path}")

    # Should not raise any errors
    mentions = read_product_mentions(report_path)
    assert isinstance(mentions, list)

    if mentions:
        # Validate structure
        assert hasattr(mentions[0], "exact_product_text")
        assert hasattr(mentions[0], "product_category")
        assert hasattr(mentions[0], "properties")
        assert hasattr(mentions[0], "email_subject")

    # Test loading complete report
    report = load_report(report_path)
    assert report["total_products"] == len(mentions)
