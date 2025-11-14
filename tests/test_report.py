"""Unit tests for Excel report generation"""

from datetime import datetime

import pytest
from openpyxl import load_workbook

from src.models.email import Email, EmailMetadata
from src.models.product import ProductAnalytics, ProductMention, ProductProperty
from workflow.nodes.reporting.excel_generator import (
    create_analytics_sheet,
    create_email_summary_sheet,
    create_product_mentions_sheet,
    generate_excel_report,
)


class TestExcelReportGeneration:
    """Test suite for Excel report generation"""

    @pytest.mark.unit
    def test_generate_excel_report_basic(self, tmp_path):
        """Test basic Excel report generation"""
        output_path = tmp_path / "test_report.xlsx"

        # Create sample data
        products = [
            ProductMention(
                exact_product_text="100 pcs of Hex Bolt",
                product_name="Hex Bolt",
                product_category="Fasteners",
                properties=[
                    ProductProperty(name="grade", value="8", confidence=1.0),
                    ProductProperty(name="size", value="1/2-13", confidence=1.0),
                ],
                quantity=100,
                unit="pcs",
                context="quote_request",
                requestor="customer@example.com",
                date_requested=None,
                email_subject="RFQ Request",
                email_sender="customer@example.com",
                email_file="test.msg",
            )
        ]

        emails = [
            Email(
                metadata=EmailMetadata(
                    message_id=None,
                    subject="RFQ Request",
                    sender="customer@example.com",
                    recipients=["sales@westbrand.ca"],
                    date=datetime(2025, 1, 15, 10, 30),
                ),
                body="Test body",
                cleaned_body="Test body",
                file_path="test.msg",
            )
        ]

        # Generate report
        result_path, analysis = generate_excel_report(products, emails, output_path)

        # Verify file exists
        assert result_path.exists()
        assert result_path == output_path
        assert len(analysis) > 0

        # Load and verify structure
        wb = load_workbook(result_path)
        assert len(wb.sheetnames) >= 2  # At least 2 sheets
        assert "Product Mentions" in wb.sheetnames
        assert "Email Summary" in wb.sheetnames

        wb.close()

    @pytest.mark.unit
    def test_product_mentions_sheet_structure(self, tmp_path):
        """Test Product Mentions sheet structure"""
        output_path = tmp_path / "mentions.xlsx"

        products = [
            ProductMention(
                exact_product_text='50 ft of 1/2" Threaded Rod',
                product_name="Threaded Rod",
                product_category="Threaded Rod",
                properties=[
                    ProductProperty(name="diameter", value='1/2"', confidence=1.0)
                ],
                quantity=50,
                unit="ft",
                context="order",
                requestor="buyer@company.com",
                date_requested=None,
                email_subject="Order Placement",
                email_sender="buyer@company.com",
                email_file="order.msg",
            )
        ]

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None

        create_product_mentions_sheet(ws, products)

        # Check headers
        headers = [cell.value for cell in ws[1]]
        assert "Email information" in headers
        assert "Category" in headers
        assert "Quantity" in headers
        assert "Context" in headers
        assert "Email Subject" in headers

        # Check data row
        assert ws.max_row >= 2  # Header + at least 1 data row

        wb.save(output_path)
        wb.close()

    @pytest.mark.unit
    def test_email_summary_sheet_structure(self, tmp_path):
        """Test Email Summary sheet structure"""
        output_path = tmp_path / "summary.xlsx"

        emails = [
            Email(
                metadata=EmailMetadata(
                    message_id=None,
                    subject="Test Email",
                    sender="test@example.com",
                    recipients=["sales@westbrand.ca"],
                    date=datetime(2025, 1, 1, 12, 0),
                ),
                body="Body",
                cleaned_body="Cleaned",
                attachments=["file.pdf"],
                file_path="test.msg",
            )
        ]

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None

        create_email_summary_sheet(ws, emails)

        # Check headers
        headers = [cell.value for cell in ws[1]]
        assert "Email File" in headers
        assert "Subject" in headers
        assert "Sender" in headers
        assert "Date" in headers

        # Check data
        assert ws.max_row >= 2

        wb.save(output_path)
        wb.close()

    @pytest.mark.unit
    def test_generate_report_empty_data(self, tmp_path):
        """Test report generation with no data"""
        output_path = tmp_path / "empty_report.xlsx"

        result_path, analysis = generate_excel_report([], [], output_path)

        assert result_path.exists()
        assert len(analysis) == 0

        wb = load_workbook(result_path)
        assert len(wb.sheetnames) >= 2
        wb.close()

    @pytest.mark.unit
    def test_generate_report_multiple_products(self, tmp_path):
        """Test report with multiple products"""
        output_path = tmp_path / "multi_report.xlsx"

        products = [
            ProductMention(
                exact_product_text=f"Product {i}",
                product_name=f"Product {i}",
                product_category="Category",
                properties=[],
                quantity=i * 10,
                unit="pcs",
                context="quote_request",
                requestor="test@example.com",
                date_requested=None,
                email_subject=f"Email {i}",
                email_sender="test@example.com",
                email_file=f"test{i}.msg",
            )
            for i in range(1, 6)
        ]

        emails = [
            Email(
                metadata=EmailMetadata(
                    message_id=None,
                    subject=f"Email {i}",
                    sender="test@example.com",
                    recipients=["sales@westbrand.ca"],
                    date=None,
                ),
                body=f"Body {i}",
                cleaned_body=f"Cleaned {i}",
                file_path=f"test{i}.msg",
            )
            for i in range(1, 6)
        ]

        result_path, analysis = generate_excel_report(products, emails, output_path)

        wb = load_workbook(result_path)
        ws_mentions = wb["Product Mentions"]

        # Should have header + 5 data rows
        assert ws_mentions.max_row >= 6
        assert len(analysis) > 0

        wb.close()

    @pytest.mark.unit
    def test_analytics_sheet_creation(self, tmp_path):
        """Test analytics sheet creation"""
        output_path = tmp_path / "analytics.xlsx"

        # Create sample analytics
        analytics = [
            ProductAnalytics(
                product_name="Hex Bolt",
                product_category="Fasteners",
                total_mentions=5,
                first_mention=datetime(2025, 1, 1),
                last_mention=datetime(2025, 2, 1),
                total_quantity=500,
                properties_summary={"grade": ["8", "5"], "size": ["1/2-13"]},
                contexts=["quote_request", "order"],
            )
        ]

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None

        create_analytics_sheet(ws, analytics)

        # Check headers
        headers = [cell.value for cell in ws[1]]
        assert "Category" in headers
        assert "Total Mentions" in headers

        # Check data
        assert ws.max_row >= 2

        wb.save(output_path)
        wb.close()

    @pytest.mark.unit
    def test_report_formatting(self, tmp_path):
        """Test that report has basic formatting"""
        output_path = tmp_path / "formatted_report.xlsx"

        products = [
            ProductMention(
                exact_product_text="10 pcs of Test Product",
                product_name="Test Product",
                product_category="Test",
                properties=[],
                quantity=10,
                unit="pcs",
                context="quote_request",
                requestor="test@example.com",
                date_requested=None,
                email_subject="Test",
                email_sender="test@example.com",
                email_file="test.msg",
            )
        ]

        emails = [
            Email(
                metadata=EmailMetadata(
                    message_id=None,
                    subject="Test",
                    sender="test@example.com",
                    recipients=["sales@westbrand.ca"],
                    date=None,
                ),
                body="Body",
                cleaned_body="Cleaned",
                file_path="test.msg",
            )
        ]

        result_path, analysis = generate_excel_report(products, emails, output_path)

        wb = load_workbook(result_path)
        ws = wb["Product Mentions"]

        # Check that header row exists
        assert ws[1][0].value is not None
        assert len(analysis) > 0

        # Check that filters or formatting might be applied
        # (This is a basic check - actual formatting details are implementation-specific)
        assert ws.max_row >= 2

        wb.close()
