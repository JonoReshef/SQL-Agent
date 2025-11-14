#!/usr/bin/env python3
"""Analyze detected properties from Excel report"""

from openpyxl import load_workbook
from collections import defaultdict

wb = load_workbook("output/report_20251113_154940.xlsx")
ws = wb["Product Mentions"]

# Extract all properties from the Properties column (column 3)
properties_by_product = defaultdict(set)

for row in ws.iter_rows(min_row=2, values_only=True):
    product_name = row[0]
    properties_str = row[2]

    if properties_str:
        prop_pairs = properties_str.split(", ")  # type: ignore
        for pair in prop_pairs:
            if "=" in pair:
                prop_name, prop_value = pair.split("=", 1)
                properties_by_product[product_name].add(
                    (prop_name.strip(), prop_value.strip())
                )

# Print findings
for product, props in sorted(properties_by_product.items()):
    print(f"\n{product}:")
    prop_dict = defaultdict(list)
    for prop_name, prop_value in sorted(props):
        prop_dict[prop_name].append(prop_value)

    for prop_name, values in sorted(prop_dict.items()):
        print(f"  {prop_name}:")
        for val in sorted(set(values))[:15]:  # Show up to 15 unique examples
            print(f"    - {repr(val)}")
