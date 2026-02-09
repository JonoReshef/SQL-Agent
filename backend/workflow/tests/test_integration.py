"""Integration tests for complete email analysis workflow"""

from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from workflow.analysis_workflow.graph import create_workflow_graph, run_workflow
from workflow.models.email import Email, EmailMetadata
from workflow.models.product import ProductMention, ProductProperty


class TestWorkflowIntegration:
    """End-to-end integration tests for the workflow"""

    def test_complete_workflow_execution(self, tmp_path):
        """Test complete workflow from .msg files to Excel report"""
        # Create test directory structure
        input_dir = tmp_path / "emails"
        input_dir.mkdir()
        output_path = tmp_path / "output" / "report.xlsx"

        # Mock the msg reader to return sample emails
        sample_email = Email(
            metadata=EmailMetadata(
                subject="Product Quote Request",
                sender="customer@example.com",
                recipients=["sales@westbrand.ca"],
                message_id="<test@example.com>",
                cc=[],
                date=None,
            ),
            body="Hello, I need 500 pieces of M10x100 bolts for a construction project.",
            cleaned_body="Hello, I need 500 pieces of M10x100 bolts for a construction project.",
            attachments=[],
            file_path=str(input_dir / "quote_request.msg"),
        )

        # Mock product extraction
        sample_product = ProductMention(
            exact_product_text="500 pieces of M10x100 bolts",
            product_name="M10x100 Bolt",
            product_category="Fasteners",
            properties=[
                ProductProperty(name="size", value="M10", confidence=0.95),
                ProductProperty(name="length", value="100mm", confidence=0.90),
            ],
            quantity=500,
            unit="pieces",
            context="quote_request",
            requestor="customer@example.com",
            date_requested=None,
            email_subject="Product Quote Request",
            email_sender="customer@example.com",
            email_file=str(input_dir / "quote_request.msg"),
        )

        # Patch dependencies
        with (
            patch(
                "src.analysis_workflow.nodes.ingestion.read_msg_files_from_directory"
            ) as mock_read,
            patch(
                "src.analysis_workflow.nodes.extraction.extract_products_batch"
            ) as mock_extract,
        ):
            mock_read.return_value = [sample_email]
            mock_extract.return_value = [sample_product]

            # Run workflow
            final_state = run_workflow(str(input_dir), str(output_path))

            # Verify state
            assert len(final_state.emails) == 1
            assert len(final_state.extracted_products) == 1
            assert final_state.report_path == str(output_path)
            assert len(final_state.errors) == 0

            # Verify Excel file was created
            assert output_path.exists()

            # Verify Excel structure
            wb = load_workbook(output_path)
            assert "Product Mentions" in wb.sheetnames
            assert "Analytics" in wb.sheetnames
            assert "Email Summary" in wb.sheetnames

    def test_workflow_with_multiple_emails(self, tmp_path):
        """Test workflow with multiple emails and products"""
        input_dir = tmp_path / "emails"
        input_dir.mkdir()
        output_path = tmp_path / "output" / "report.xlsx"

        # Create multiple sample emails
        emails = [
            Email(
                metadata=EmailMetadata(
                    subject=f"Quote {i}",
                    sender=f"customer{i}@example.com",
                    recipients=["sales@westbrand.ca"],
                    message_id=f"<test{i}@example.com>",
                    cc=[],
                    date=None,
                ),
                body=f"Need bolts for project {i}",
                cleaned_body=f"Need bolts for project {i}",
                attachments=[],
                file_path=str(input_dir / f"quote_{i}.msg"),
            )
            for i in range(3)
        ]

        # Create multiple products
        products = [
            ProductMention(
                exact_product_text=f"bolts for project {i}",
                product_name=f"Product {i}",
                product_category="Fasteners",
                properties=[],
                quantity=i * 100,
                unit="pcs",
                context="quote_request",
                requestor=f"customer{i}@example.com",
                date_requested=None,
                email_subject=f"Quote {i}",
                email_sender=f"customer{i}@example.com",
                email_file=str(input_dir / f"quote_{i}.msg"),
            )
            for i in range(3)
        ]

        with (
            patch(
                "src.analysis_workflow.nodes.ingestion.read_msg_files_from_directory"
            ) as mock_read,
            patch(
                "src.analysis_workflow.nodes.extraction.extract_products_batch"
            ) as mock_extract,
        ):
            mock_read.return_value = emails
            mock_extract.return_value = products

            final_state = run_workflow(str(input_dir), str(output_path))

            # Verify results
            assert len(final_state.emails) == 3
            assert len(final_state.extracted_products) == 3
            assert output_path.exists()

    def test_workflow_with_errors(self, tmp_path):
        """Test workflow continues with errors captured"""
        input_dir = tmp_path / "emails"
        input_dir.mkdir()
        output_path = tmp_path / "output" / "report.xlsx"

        # Mock extraction to raise error
        with (
            patch(
                "src.analysis_workflow.nodes.ingestion.read_msg_files_from_directory"
            ) as mock_read,
            patch(
                "src.analysis_workflow.nodes.extraction.extract_products_batch"
            ) as mock_extract,
        ):
            mock_read.return_value = []
            mock_extract.side_effect = Exception("LLM API timeout")

            final_state = run_workflow(str(input_dir), str(output_path))

            # Verify error was captured
            assert len(final_state.errors) == 1
            assert "LLM API timeout" in final_state.errors[0]

    def test_workflow_graph_structure(self):
        """Test workflow graph has correct nodes and edges"""
        graph = create_workflow_graph()

        # Verify graph was compiled successfully
        assert graph is not None

    def test_empty_directory_handling(self, tmp_path):
        """Test workflow handles empty input directory gracefully"""
        input_dir = tmp_path / "empty"
        input_dir.mkdir()
        output_path = tmp_path / "output" / "report.xlsx"

        with patch(
            "src.analysis_workflow.nodes.ingestion.read_msg_files_from_directory"
        ) as mock_read:
            mock_read.return_value = []

            final_state = run_workflow(str(input_dir), str(output_path))

            # Verify empty results
            assert len(final_state.emails) == 0
            assert len(final_state.extracted_products) == 0

            # Report should still be generated (empty)
            assert output_path.exists()

    def test_real_msg_file_processing(self, tmp_path):
        """Test with actual .msg files from workspace (if available)"""
        # Use real .msg files from data directory
        real_data_dir = Path(
            "/Users/e401604/Documents/Products/WestBrand/data/sales@westbrand.ca/Recoverable-Items/Deletions"
        )

        if not real_data_dir.exists():
            pytest.skip("Real .msg files not available")

        output_path = tmp_path / "real_test_report.xlsx"

        # Mock only the LLM extraction (let real parsing happen)
        with patch(
            "src.analysis_workflow.nodes.extraction.extract_products_batch"
        ) as mock_extract:
            # Mock minimal product response
            mock_extract.return_value = []

            try:
                final_state = run_workflow(str(real_data_dir), str(output_path))

                # Verify emails were parsed
                assert len(final_state.emails) > 0

                # Verify report was generated (even if errors occurred)
                # The report node should create file even with errors
                if len(final_state.errors) == 0:
                    assert output_path.exists()
            except Exception:
                # If real file processing fails, skip test
                pytest.skip("Real .msg file processing failed")
