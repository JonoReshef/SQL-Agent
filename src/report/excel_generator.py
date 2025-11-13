"""Excel report generation for product analysis"""

from datetime import datetime
from pathlib import Path
from typing import List
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from src.models.product import ProductMention, ProductAnalytics
from src.models.email import Email


def sanitize_for_excel(value):
    """
    Sanitize values for Excel compatibility.

    - Removes illegal characters from strings
    - Strips timezone info from datetime objects

    Args:
        value: Value to sanitize

    Returns:
        Sanitized value safe for Excel cells
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        # Excel doesn't support timezone-aware datetimes
        return value.replace(tzinfo=None)
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def generate_excel_report(
    products: List[ProductMention], emails: List[Email], output_path: Path
) -> Path:
    """
    Generate comprehensive Excel report.

    Args:
        products: List of product mentions
        emails: List of emails processed
        output_path: Path where to save the Excel file

    Returns:
        Path to generated Excel file
    """
    # Create workbook
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Create sheets
    ws_mentions = wb.create_sheet("Product Mentions")
    create_product_mentions_sheet(ws_mentions, products)

    # Create analytics if we have products
    if products:
        analytics = calculate_analytics(products)
        ws_analytics = wb.create_sheet("Analytics")
        create_analytics_sheet(ws_analytics, analytics)

    # Create email summary
    ws_summary = wb.create_sheet("Email Summary")
    create_email_summary_sheet(ws_summary, emails)

    # Save workbook
    wb.save(output_path)
    wb.close()

    return output_path


def create_product_mentions_sheet(
    ws: Worksheet, products: List[ProductMention]
) -> None:
    """
    Create Product Mentions sheet with all product details.

    Args:
        ws: Worksheet to populate
        products: List of product mentions
    """
    # Define headers
    headers = [
        "Product extracted",
        "Category",
        "Properties",
        "Requestor",
        "Quantity",
        "Unit",
        "Context",
        "Date Requested",
        "Email Subject",
        "Sender",
        "Email Date",
        "File",
    ]

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row_idx, product in enumerate(products, start=2):
        # Format properties as string
        props_str = ", ".join([f"{p.name}={p.value}" for p in product.properties])

        ws.cell(
            row=row_idx, column=1, value=sanitize_for_excel(product.exact_product_text)
        )
        ws.cell(
            row=row_idx, column=2, value=sanitize_for_excel(product.product_category)
        )

        ws.cell(row=row_idx, column=3, value=sanitize_for_excel(props_str))
        ws.cell(row=row_idx, column=4, value=sanitize_for_excel(product.requestor))
        ws.cell(row=row_idx, column=5, value=product.quantity)
        ws.cell(row=row_idx, column=6, value=sanitize_for_excel(product.unit))
        ws.cell(row=row_idx, column=7, value=sanitize_for_excel(product.context))
        ws.cell(row=row_idx, column=8, value=sanitize_for_excel(product.date_requested))
        ws.cell(row=row_idx, column=9, value=sanitize_for_excel(product.email_subject))
        ws.cell(row=row_idx, column=10, value=sanitize_for_excel(product.email_sender))
        ws.cell(row=row_idx, column=11, value=sanitize_for_excel(product.email_date))
        ws.cell(row=row_idx, column=12, value=sanitize_for_excel(product.email_file))
    # Auto-fit columns
    for col_idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add auto-filter
    if products:
        ws.auto_filter.ref = f"A1:K{len(products) + 1}"


def create_analytics_sheet(ws: Worksheet, analytics: List[ProductAnalytics]) -> None:
    """
    Create Analytics sheet with aggregated data.

    Args:
        ws: Worksheet to populate
        analytics: List of product analytics
    """
    # Define headers
    headers = [
        "Product",
        "Category",
        "Total Mentions",
        "First Mention",
        "Last Mention",
        "Total Quantity",
        "Property Variations",
        "Contexts",
    ]

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="FFE5CC", end_color="FFE5CC", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row_idx, analytic in enumerate(analytics, start=2):
        # Format property variations
        props_str = "; ".join(
            [
                f"{key}: {', '.join(values)}"
                for key, values in analytic.properties_summary.items()
            ]
        )

        # Format contexts
        contexts_str = ", ".join(analytic.contexts)

        ws.cell(row=row_idx, column=1, value=sanitize_for_excel(analytic.product_name))
        ws.cell(
            row=row_idx, column=2, value=sanitize_for_excel(analytic.product_category)
        )
        ws.cell(row=row_idx, column=3, value=analytic.total_mentions)
        ws.cell(row=row_idx, column=4, value=sanitize_for_excel(analytic.first_mention))
        ws.cell(row=row_idx, column=5, value=sanitize_for_excel(analytic.last_mention))
        ws.cell(row=row_idx, column=6, value=analytic.total_quantity)
        ws.cell(row=row_idx, column=7, value=sanitize_for_excel(props_str))
        ws.cell(row=row_idx, column=8, value=sanitize_for_excel(contexts_str))

    # Auto-fit columns
    for col_idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"


def create_email_summary_sheet(ws: Worksheet, emails: List[Email]) -> None:
    """
    Create Email Summary sheet.

    Args:
        ws: Worksheet to populate
        emails: List of emails processed
    """
    # Define headers
    headers = [
        "Email File",
        "Subject",
        "Sender",
        "Recipients",
        "Date",
        "Has Attachments",
        "Body Length",
    ]

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="E5FFCC", end_color="E5FFCC", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row_idx, email in enumerate(emails, start=2):
        recipients_str = ", ".join(email.metadata.recipients)
        has_attachments = "Yes" if email.attachments else "No"
        body_length = len(email.cleaned_body) if email.cleaned_body else len(email.body)

        ws.cell(row=row_idx, column=1, value=sanitize_for_excel(email.file_path))
        ws.cell(row=row_idx, column=2, value=sanitize_for_excel(email.metadata.subject))
        ws.cell(row=row_idx, column=3, value=sanitize_for_excel(email.metadata.sender))
        ws.cell(row=row_idx, column=4, value=sanitize_for_excel(recipients_str))
        ws.cell(row=row_idx, column=5, value=sanitize_for_excel(email.metadata.date))
        ws.cell(row=row_idx, column=6, value=has_attachments)
        ws.cell(row=row_idx, column=7, value=body_length)

    # Auto-fit columns
    for col_idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"


def calculate_analytics(products: List[ProductMention]) -> List[ProductAnalytics]:
    """
    Calculate aggregated analytics from product mentions.

    Args:
        products: List of product mentions

    Returns:
        List of product analytics
    """
    # Group by product name
    product_groups = {}

    for product in products:
        key = (product.product_name, product.product_category)
        if key not in product_groups:
            product_groups[key] = []
        product_groups[key].append(product)

    # Calculate analytics for each group
    analytics = []

    for (product_name, category), mentions in product_groups.items():
        # Get dates
        dates = [m.email_date for m in mentions if m.email_date]
        first_mention = min(dates) if dates else None
        last_mention = max(dates) if dates else None

        # Total quantity
        total_quantity = sum(m.quantity for m in mentions if m.quantity)

        # Property variations
        properties_summary = {}
        for mention in mentions:
            for prop in mention.properties:
                if prop.name not in properties_summary:
                    properties_summary[prop.name] = []
                if prop.value not in properties_summary[prop.name]:
                    properties_summary[prop.name].append(prop.value)

        # Unique contexts
        contexts = list(set(m.context for m in mentions))

        analytic = ProductAnalytics(
            product_name=product_name,
            product_category=category,
            total_mentions=len(mentions),
            first_mention=first_mention,
            last_mention=last_mention,
            total_quantity=total_quantity if total_quantity > 0 else None,
            properties_summary=properties_summary,
            contexts=contexts,
        )

        analytics.append(analytic)

    # Sort by total mentions (descending)
    analytics.sort(key=lambda x: x.total_mentions, reverse=True)

    return analytics
