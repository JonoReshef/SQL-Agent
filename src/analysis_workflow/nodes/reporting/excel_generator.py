"""Excel report generation for product analysis"""

from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.models.email import Email
from src.models.inventory import InventoryMatch, ReviewFlag
from src.models.product import ProductAnalytics, ProductMention
from src.utils.compute_content_hash import compute_content_hash


def generate_excel_report(
    all_products: List[ProductMention],
    unique_property_products: List[ProductMention],
    emails: List[Email],
    output_path: Path,
    product_matches: dict[str, List[InventoryMatch]] | None = None,
    review_flags: List[ReviewFlag] | None = None,
) -> tuple[Path, List[ProductAnalytics]]:
    """
    Generate comprehensive Excel report.

    Args:
        products: List of product mentions
        emails: List of emails processed
        output_path: Path where to save the Excel file
        product_matches: Dictionary mapping product text to list of inventory matches
        review_flags: List of review flags for products needing manual review

    Returns:
        Path to generated Excel file and analytics
    """
    # Create workbook
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Create sheets
    ws_mentions = wb.create_sheet("Product Mentions")
    create_product_mentions_sheet(ws_mentions, all_products)

    # Create analytics if we have products
    if all_products:
        analytics = _calculate_analytics(all_products)
        ws_analytics = wb.create_sheet("Analytics")
        create_analytics_sheet(ws_analytics, analytics)
    else:
        analytics = []

    # Create email summary
    ws_summary = wb.create_sheet("Email Summary")
    create_email_summary_sheet(ws_summary, emails, all_products)

    # Create inventory matches sheet if matches provided
    if product_matches:
        ws_matches = wb.create_sheet("Inventory Matches")
        _create_inventory_matches_sheet(ws_matches, unique_property_products, product_matches)

    # Create review flags sheet if flags provided
    if review_flags:
        ws_flags = wb.create_sheet("Review Flags")
        _create_review_flags_sheet(ws_flags, review_flags)

    # Save workbook
    wb.save(output_path)
    wb.close()

    return output_path, analytics


def create_product_mentions_sheet(ws: Worksheet, products: List[ProductMention]) -> None:
    """
    Create Product Mentions sheet with all product details.

    Args:
        ws: Worksheet to populate
        products: List of product mentions
    """
    # Define headers
    headers = [
        "Email information",
        "Category",
        "Properties",
        "Requestor",
        "Quantity",
        "Unit",
        "Context",
        "Date Requested",
        "Email Subject",
        "Sender",
        "File",
    ]

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row_idx, product in enumerate(products, start=2):
        # Format properties as string
        props_str = ", ".join([f"{p.name}={p.value}" for p in product.properties])

        ws.cell(row=row_idx, column=1, value=_sanitize_for_excel(product.exact_product_text))
        ws.cell(row=row_idx, column=2, value=_sanitize_for_excel(product.product_category))

        ws.cell(row=row_idx, column=3, value=_sanitize_for_excel(props_str))
        ws.cell(row=row_idx, column=4, value=_sanitize_for_excel(product.requestor))
        ws.cell(row=row_idx, column=5, value=product.quantity)
        ws.cell(row=row_idx, column=6, value=_sanitize_for_excel(product.unit))
        ws.cell(row=row_idx, column=7, value=_sanitize_for_excel(product.context))
        ws.cell(row=row_idx, column=8, value=_sanitize_for_excel(product.date_requested))
        ws.cell(row=row_idx, column=9, value=_sanitize_for_excel(product.email_subject))
        ws.cell(row=row_idx, column=10, value=_sanitize_for_excel(product.requestor))
        ws.cell(row=row_idx, column=12, value=_sanitize_for_excel(product.email_file))
    # Auto-fit columns
    for col_idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add auto-filter
    if products:
        ws.auto_filter.ref = f"A1:K{len(products) + 1}"


def create_analytics_sheet(ws: Worksheet, analytics: List[ProductAnalytics]) -> None:
    """
    Create Analytics sheet with aggregated data.

    Args:
        ws: Worksheet to populate
        analytics: List of product analytics
    """
    # Define headers
    headers = [
        "Product Name",
        "Category",
        "Total Mentions",
        "First Mention",
        "Last Mention",
        "Total Quantity",
        "Property Variations",
        "Contexts",
        "People Involved",
    ]

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row_idx, analytic in enumerate(analytics, start=2):
        # Format property variations
        props_str = "; ".join(
            [f"{key}: {', '.join(values)}" for key, values in analytic.properties_summary.items()]
        )

        # Format contexts
        contexts_str = ", ".join(analytic.contexts)

        ws.cell(row=row_idx, column=1, value=_sanitize_for_excel(analytic.product_name))
        ws.cell(row=row_idx, column=2, value=_sanitize_for_excel(analytic.product_category))
        ws.cell(row=row_idx, column=3, value=analytic.total_mentions)
        ws.cell(row=row_idx, column=4, value=_sanitize_for_excel(analytic.first_mention))
        ws.cell(row=row_idx, column=5, value=_sanitize_for_excel(analytic.last_mention))
        ws.cell(row=row_idx, column=6, value=analytic.total_quantity)
        ws.cell(row=row_idx, column=7, value=_sanitize_for_excel(props_str))
        ws.cell(row=row_idx, column=8, value=_sanitize_for_excel(contexts_str))
        ws.cell(
            row=row_idx,
            column=9,
            value=_sanitize_for_excel(", ".join(analytic.people_involved)),
        )

    # Auto-fit columns
    for col_idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"


def create_email_summary_sheet(
    ws: Worksheet, emails: List[Email], products: List[ProductMention]
) -> None:
    """
    Create Email Summary sheet.

    Args:
        ws: Worksheet to populate
        emails: List of emails processed
    """
    # Define headers
    headers = [
        "Email File",
        "Subject",
        "Sender",
        "Recipients",
        "Date",
        "Number of Products Mentioned",
        "Has Attachments",
        "Body Length (words)",
    ]

    # Map email thread_hash to number of products mentioned
    email_product_count = {email.thread_hash: 0 for email in emails}
    for product in products:
        if product.thread_hash in email_product_count:
            email_product_count[product.thread_hash] += 1

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E5FFCC", end_color="E5FFCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row_idx, email in enumerate(emails, start=2):
        recipients_str = ", ".join(email.metadata.recipients)
        has_attachments = "Yes" if email.attachments else "No"
        body_length = len(email.cleaned_body.split()) if email.cleaned_body else "Unknown"

        ws.cell(row=row_idx, column=1, value=_sanitize_for_excel(email.file_path))
        ws.cell(row=row_idx, column=2, value=_sanitize_for_excel(email.metadata.subject))
        ws.cell(row=row_idx, column=3, value=_sanitize_for_excel(email.metadata.sender))
        ws.cell(row=row_idx, column=4, value=_sanitize_for_excel(recipients_str))
        ws.cell(row=row_idx, column=5, value=_sanitize_for_excel(email.metadata.date))
        ws.cell(row=row_idx, column=6, value=email_product_count.get(email.thread_hash, 0))
        ws.cell(row=row_idx, column=7, value=has_attachments)
        ws.cell(row=row_idx, column=8, value=body_length)

    # Auto-fit columns
    for col_idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"


def _calculate_analytics(products: List[ProductMention]) -> List[ProductAnalytics]:
    """
    Calculate aggregated analytics from product mentions.

    Args:
        products: List of product mentions

    Returns:
        List of product analytics
    """
    # Group by product name
    product_groups = {}

    for product in products:
        key = (product.product_name, product.product_category)
        if key not in product_groups:
            product_groups[key] = []
        product_groups[key].append(product)

    # Calculate analytics for each group
    analytics = []

    for (product_name, category), mentions in product_groups.items():
        # Get dates
        dates = [
            m.date_requested for m in mentions if m.date_requested
        ]  # date_requested is a string but formatted as yyyy-MM-dd or yyyy-MM-ddTHH:mm:ss so we can compare them lexicographically
        first_mention = min(dates) if dates else None
        last_mention = max(dates) if dates else None

        # Total quantity
        total_quantity = sum(m.quantity for m in mentions if m.quantity)

        # People involved
        people = list(set(m.requestor for m in mentions if m.requestor))

        # Property variations
        properties_summary = {}
        for mention in mentions:
            for prop in mention.properties:
                if prop.name not in properties_summary:
                    properties_summary[prop.name] = []
                if prop.value not in properties_summary[prop.name]:
                    properties_summary[prop.name].append(prop.value)

        # Unique contexts
        contexts = list(set(m.context for m in mentions))

        analytic = ProductAnalytics(
            product_name=product_name,
            product_category=category,
            total_mentions=len(mentions),
            first_mention=first_mention,
            last_mention=last_mention,
            total_quantity=total_quantity if total_quantity > 0 else None,
            properties_summary=properties_summary,
            people_involved=people,
            contexts=contexts,
        )

        analytics.append(analytic)

    # Sort by total mentions (descending)
    analytics.sort(key=lambda x: x.total_mentions, reverse=True)

    return analytics


def _write_product_info(row_idx: int, ws: Worksheet, product: ProductMention) -> int:
    """
    Helper function which writes the product info to the current row
    """
    col_idx = 1
    ws.cell(  # Product Text
        row=row_idx,
        column=col_idx,
        value=_sanitize_for_excel(product.exact_product_text),
    )
    col_idx += 1
    ws.cell(  # Product Name
        row=row_idx,
        column=col_idx,
        value=_sanitize_for_excel(product.product_name),
    )
    col_idx += 1
    ws.cell(  # Category
        row=row_idx,
        column=col_idx,
        value=_sanitize_for_excel(product.product_category),
    )
    col_idx += 1
    ws.cell(  # Product Properties
        row=row_idx,
        column=col_idx,
        value=_sanitize_for_excel(
            "; ".join([f"{value.name}: {value.value}" for value in product.properties])
        ),
    )
    col_idx += 1
    ws.cell(  # Email Subject
        row=row_idx,
        column=col_idx,
        value=_sanitize_for_excel(product.email_subject),
    )
    col_idx += 1
    ws.cell(  # Email Sender
        row=row_idx,
        column=col_idx,
        value=_sanitize_for_excel(product.email_sender),
    )
    col_idx += 1
    return col_idx


def _create_inventory_matches_sheet(
    ws: Worksheet,
    products: List[ProductMention],
    product_matches: dict[str, List[InventoryMatch]],
) -> None:
    """
    Create Inventory Matches sheet showing all product-to-inventory matches.

    Args:
        ws: Worksheet to populate
        products: List of product mentions
        product_matches: Dictionary mapping product text to inventory matches
    """
    # Define headers
    headers = [
        "Product Text",
        "Product Name",
        "Category",
        "Product Properties",
        "Email Subject",
        "Sender",
        "Inventory Item #",
        "Inventory Description",
        "Number of matched properties",
        "Inventory Properties",
        "Match Score",
        "Rank",
        "Matched Properties",
        "Missing Properties",
        "Match Reasoning",
    ]
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D0F0C0", end_color="D0F0C0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    row_idx = 2
    for product in products:
        matches = product_matches.get(compute_content_hash(product), [])

        # Show product with no matches
        if not matches:
            col_idx = _write_product_info(row_idx, ws, product)
            ws.cell(  # No Matches
                row=row_idx,
                column=col_idx,
                value="NO MATCHES",
            )
            # Highlight no match rows in light red
            for col in range(1, 16):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color="FFD7D7", end_color="FFD7D7", fill_type="solid"
                )
            row_idx += 1
        else:
            # Show each match as a separate row
            for match in matches:
                col_idx = _write_product_info(row_idx, ws, product)
                ws.cell(  # Inventory Item #
                    row=row_idx,
                    column=col_idx,
                    value=_sanitize_for_excel(match.inventory_item_number),
                )
                col_idx += 1
                ws.cell(  # Inventory Description
                    row=row_idx,
                    column=col_idx,
                    value=_sanitize_for_excel(match.inventory_description),
                )
                col_idx += 1
                ws.cell(  # Number of matched properties
                    row=row_idx,
                    column=col_idx,
                    value=len(match.matched_properties),
                )
                col_idx += 1
                ws.cell(  # Inventory Properties
                    row=row_idx,
                    column=col_idx,
                    value=_sanitize_for_excel(
                        "; ".join(
                            [f"{prop.name}: {prop.value}" for prop in match.inventory_properties]
                        )
                    ),
                )
                col_idx += 1
                ws.cell(  # Match Score
                    row=row_idx,
                    column=col_idx,
                    value=match.match_score,
                )
                # Color code based on match score
                score_cell = ws.cell(row=row_idx, column=col_idx)
                if match.match_score >= 0.8:
                    # High confidence - green
                    score_cell.fill = PatternFill(
                        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                    )
                elif match.match_score >= 0.6:
                    # Medium confidence - yellow
                    score_cell.fill = PatternFill(
                        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                    )
                else:
                    # Low confidence - orange
                    score_cell.fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )

                col_idx += 1
                ws.cell(  # Rank
                    row=row_idx,
                    column=col_idx,
                    value=match.rank
                    if len(matches) > 1
                    else "",  # If there are multiple matches, show rank
                )
                col_idx += 1
                ws.cell(  # Matched Properties
                    row=row_idx,
                    column=col_idx,
                    value=_sanitize_for_excel(", ".join(match.matched_properties)),
                )
                col_idx += 1
                ws.cell(  # Missing Properties
                    row=row_idx,
                    column=col_idx,
                    value=_sanitize_for_excel(", ".join(match.missing_properties)),
                )
                col_idx += 1
                ws.cell(  # Match Reasoning
                    row=row_idx,
                    column=col_idx,
                    value=_sanitize_for_excel(match.match_reasoning),
                )

                row_idx += 1

    # Auto-fit columns
    for col_idx in range(1, 15):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in ws[column]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add auto-filter
    if row_idx > 2:
        ws.auto_filter.ref = f"A1:L{row_idx - 1}"


def _create_review_flags_sheet(ws: Worksheet, review_flags: List[ReviewFlag]) -> None:
    """
    Create Review Flags sheet showing products requiring manual review.

    Args:
        ws: Worksheet to populate
        review_flags: List of review flags
    """
    # Define headers
    headers = [
        "Product Text",
        "Product Name",
        "Category",
        "Issue Type",
        "Match Count",
        "Top Confidence",
        "Reason",
        "Action Needed",
    ]

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row_idx, flag in enumerate(review_flags, start=2):
        ws.cell(row=row_idx, column=1, value=_sanitize_for_excel(flag.product_text))
        ws.cell(row=row_idx, column=2, value=_sanitize_for_excel(flag.product_name))
        ws.cell(row=row_idx, column=3, value=_sanitize_for_excel(flag.product_category))
        ws.cell(row=row_idx, column=4, value=_sanitize_for_excel(flag.issue_type))
        ws.cell(row=row_idx, column=5, value=flag.match_count)
        ws.cell(
            row=row_idx,
            column=6,
            value=flag.top_confidence if flag.top_confidence else "N/A",
        )
        ws.cell(row=row_idx, column=7, value=_sanitize_for_excel(flag.reason))
        ws.cell(row=row_idx, column=8, value=_sanitize_for_excel(flag.action_needed))

        # Color code based on issue type
        issue_cell = ws.cell(row=row_idx, column=4)
        if flag.issue_type in ["INSUFFICIENT_DATA", "AMBIGUOUS_MATCH"]:
            # High priority - red
            issue_cell.fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            )
        elif flag.issue_type == "LOW_CONFIDENCE":
            # Medium priority - yellow
            issue_cell.fill = PatternFill(
                start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
            )
        else:
            # Lower priority - light orange
            issue_cell.fill = PatternFill(
                start_color="FFE699", end_color="FFE699", fill_type="solid"
            )

    # Auto-fit columns
    for col_idx in range(1, 9):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in ws[column]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add auto-filter
    if review_flags:
        ws.auto_filter.ref = f"A1:H{len(review_flags) + 1}"


def _sanitize_for_excel(value):
    """
    Sanitize values for Excel compatibility.

    - Removes illegal characters from strings
    - Strips timezone info from datetime objects

    Args:
        value: Value to sanitize

    Returns:
        Sanitized value safe for Excel cells
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        # Excel doesn't support timezone-aware datetimes
        return value.replace(tzinfo=None).isoformat(sep=" ")
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value
